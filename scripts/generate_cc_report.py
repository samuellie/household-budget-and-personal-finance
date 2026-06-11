"""
Credit Card Commitment & Projection Report
==========================================
Analyses HSBC and AmBank CC (CARz Card Gold VISA S) transactions to identify
recurring charges and project expected expenses for the next 6 months.

Sheets:
  1. Current Status    — card balances, data coverage, monthly commitment total
  2. Recurring Items   — auto-detected recurring charges with avg amount & billing day
  3. Instalment Plans  — all active instalment plans (HSBC + AmBank CC)
  4. 6-Month Projection — recurring items projected forward month by month
  5. Transaction History — full charge history across both cards

Usage:
    python3 scripts/generate_cc_report.py
    python3 scripts/generate_cc_report.py --months 6
"""

import argparse
import csv as csv_mod
import io
import re
import subprocess
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
import config

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip3 install --break-system-packages openpyxl")
    sys.exit(1)

# ── Colours (shared palette) ──────────────────────────────────────────────────
C_HEADER_BG    = "1F3864"
C_HEADER_FG    = "FFFFFF"
C_SUBHEADER_BG = "D6E4F7"
C_ALT_ROW      = "F5F5F5"
C_TOTAL_BG     = "FFF9C4"
C_WARN_BG      = "FFCCBC"   # orange — high-spend projection cell
C_RECURRING_BG = "E8F5E9"   # green tint for confirmed recurring

MYR_FORMAT = '#,##0.00'

# Minimum number of distinct months a merchant must appear in to be flagged recurring
RECURRING_MIN_MONTHS = 2

# Merchant name cleanup — strip payment-processor prefixes and noise
_CLEAN_RE = re.compile(
    r"^(?:SALE DEBIT |PRE-AUTH DEBIT |PRE-AUTH REFUND |IPY\*|GP |FPX PAYMENT FR [A-Z]\/\s*[A-Z0-9]+\s*)",
    re.IGNORECASE,
)
_NOISE_RE = re.compile(r"\s+(KUALA LUMPUR|SELANGOR|MY|KL|SG|SGP)\s*$", re.IGNORECASE)


def clean_merchant(raw: str) -> str:
    name = _CLEAN_RE.sub("", raw.strip())
    name = _NOISE_RE.sub("", name)
    # Truncate long names
    return name[:45].strip()


# ── Ledger helpers ─────────────────────────────────────────────────────────────

def run_ledger(args: list) -> str:
    cmd = ["ledger", "-f", str(config.MAIN_JOURNAL)] + args
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode != 0:
        raise RuntimeError(f"ledger error: {result.stderr.strip()}")
    return result.stdout


def get_card_balance(account: str) -> float:
    """Return current liability balance as a positive number (amount owed)."""
    try:
        out = run_ledger(["bal", account, "--flat", "--no-total", "--exchange", "MYR"])
    except RuntimeError:
        return 0.0
    for line in out.splitlines():
        m = re.match(r"\s*(?:MYR\s+)?([-\d,]+\.\d{2})", line)
        if m:
            return abs(float(m.group(1).replace(",", "")))
    return 0.0


def get_cc_expense_postings(journal_globs: list[str]) -> list[dict]:
    """
    Parse expense postings from the given journal glob patterns.
    Returns list of dicts: date, merchant, category, amount, source.
    """
    txns = []
    for glob in journal_globs:
        source = glob.split("/")[0]  # e.g. "hsbc" or "ambcc"
        for journal in sorted(config.LEDGER_DIR.glob(glob)):
            text = journal.read_text(encoding="utf-8")
            entries = re.split(r"\n(?=\d{4}-\d{2}-\d{2})", text)
            for entry in entries:
                lines = entry.strip().splitlines()
                if not lines:
                    continue
                header_m = re.match(r"(\d{4}-\d{2}-\d{2})\s+[*!]?\s*(.*)", lines[0])
                if not header_m:
                    continue
                txn_date = datetime.strptime(header_m.group(1), "%Y-%m-%d").date()
                payee    = header_m.group(2).strip()

                category = "Unknown"
                amount   = 0.0
                for posting_line in lines[1:]:
                    posting_line = posting_line.strip()
                    if posting_line.startswith(";"):
                        continue
                    pm = re.match(r"(Expenses:[^\s]+)\s+MYR\s+([-\d,]+\.\d{2})", posting_line)
                    if pm:
                        category = pm.group(1)
                        try:
                            amount = float(pm.group(2).replace(",", ""))
                        except ValueError:
                            pass
                        break
                    if posting_line.startswith("Equity:") or posting_line.startswith("Liabilities:"):
                        amount = 0.0

                if amount > 0 and category != "Unknown":
                    txns.append({
                        "date":     txn_date,
                        "merchant": payee,
                        "category": category,
                        "amount":   amount,
                        "source":   source,
                    })

    return sorted(txns, key=lambda x: x["date"])


def detect_recurring(txns: list[dict]) -> list[dict]:
    """
    Group transactions by cleaned merchant name.
    Flag as recurring if they appear in RECURRING_MIN_MONTHS+ distinct months.
    Returns list of recurring item dicts.
    """
    # Group by cleaned merchant
    by_merchant: dict[str, list[dict]] = defaultdict(list)
    for t in txns:
        key = clean_merchant(t["merchant"])
        by_merchant[key].append(t)

    recurring = []
    for merchant, charges in by_merchant.items():
        months_seen = {(t["date"].year, t["date"].month) for t in charges}
        if len(months_seen) < RECURRING_MIN_MONTHS:
            continue

        amounts  = [t["amount"] for t in charges]
        avg_amt  = sum(amounts) / len(amounts)
        # Typical billing day = median day-of-month
        days     = sorted(t["date"].day for t in charges)
        med_day  = days[len(days) // 2]
        last_charge = max(t["date"] for t in charges)
        category = charges[-1].get("category", "Expenses:Subscriptions:Other")
        months_count = len(months_seen)

        is_sub = "subscriptions" in category.lower()
        recurring.append({
            "merchant":       merchant,
            "category":       category.replace("Expenses:", ""),
            "avg_amount":     avg_amt,
            "billing_day":    med_day,
            "months_seen":    months_count,
            "last_charge":    last_charge,
            "all_amounts":    amounts,
            "is_subscription": is_sub,
        })

    return sorted(recurring, key=lambda x: x["avg_amount"], reverse=True)


def project_months(n: int) -> list[date]:
    """Return the 1st of each of the next n months from today."""
    today = date.today()
    months = []
    y, m = today.year, today.month
    for _ in range(n):
        m += 1
        if m > 12:
            m = 1
            y += 1
        months.append(date(y, m, 1))
    return months


def _add_months(d: date, n: int) -> date:
    """Add n months to a date."""
    m = d.month - 1 + n
    y = d.year + m // 12
    m = m % 12 + 1
    return date(y, m, 1)


# HSBC format: "MERCHANT NAME - [DURATION -] X/Y INSTALMENT"
_HSBC_INSTALMENT_RE = re.compile(
    r"^(.*?)\s*-\s*(?:\d+\s*-\s*)?(\d+)/(\d+)\s+INSTAL[LM]ENT",
    re.IGNORECASE,
)

# AmBank CC format: "MERCHANT Nth/TOTAL (RMtotal)"  e.g. "LEGEND 03rd/24 (RM8,736)"
_AMBCC_INSTALMENT_RE = re.compile(
    r"^(.*?)\s+(\d+)(?:st|nd|rd|th)/(\d+)\s+\(RM[\d,]+\)",
    re.IGNORECASE,
)


def _parse_instalment_entry(payee: str, lines: list[str], txn_date: date,
                             source: str) -> dict | None:
    """
    Try to parse a journal entry as an instalment plan.
    Returns a dict or None if the entry is not an instalment.
    """
    if source == "hsbc":
        im = _HSBC_INSTALMENT_RE.match(payee)
        if not im:
            return None
        merchant  = im.group(1).strip().rstrip("-").strip()
        current_n = int(im.group(2))
        total_n   = int(im.group(3))
        # Amount from raw comment or expense posting
        amount = None
        for line in lines[1:]:
            raw_m = re.search(r"INSTAL[LM]ENT\s+([\d,]+\.\d{2})", line, re.IGNORECASE)
            if raw_m:
                try:
                    amount = float(raw_m.group(1).replace(",", ""))
                except ValueError:
                    pass
                break
            exp_m = re.match(r"\s+Expenses:[^\s]+\s+MYR\s+([\d,]+\.\d{2})", line)
            if exp_m and amount is None:
                try:
                    amount = float(exp_m.group(1).replace(",", ""))
                except ValueError:
                    pass

    elif source == "ambcc":
        im = _AMBCC_INSTALMENT_RE.match(payee)
        if not im:
            return None
        merchant  = im.group(1).strip()
        current_n = int(im.group(2))
        total_n   = int(im.group(3))
        # Amount from expense posting line
        amount = None
        for line in lines[1:]:
            exp_m = re.match(r"\s+Expenses:[^\s]+\s+MYR\s+([\d,]+\.\d{2})", line)
            if exp_m:
                try:
                    amount = float(exp_m.group(1).replace(",", ""))
                except ValueError:
                    pass
                break
    else:
        return None

    if amount is None:
        return None

    return {
        "merchant":  merchant,
        "current_n": current_n,
        "total_n":   total_n,
        "amount":    amount,
        "date":      txn_date,
        "source":    source,
    }


def extract_installments() -> list[dict]:
    """
    Parse HSBC and AmBank CC journal files for instalment plan entries.
    Returns one dict per plan (keyed on merchant+total+source), tracking
    the latest instalment number seen and remaining months.
    """
    plans: dict[tuple, dict] = {}

    sources = [
        ("hsbc",  "hsbc/*.journal"),
        ("ambcc", "ambcc/*.journal"),
    ]

    for source, glob in sources:
        for journal in sorted(config.LEDGER_DIR.glob(glob)):
            text    = journal.read_text(encoding="utf-8")
            entries = re.split(r"\n(?=\d{4}-\d{2}-\d{2})", text)

            for entry in entries:
                lines = entry.strip().splitlines()
                if not lines:
                    continue
                hm = re.match(r"(\d{4}-\d{2}-\d{2})\s+[*!]?\s*(.*)", lines[0])
                if not hm:
                    continue
                txn_date = datetime.strptime(hm.group(1), "%Y-%m-%d").date()
                payee    = hm.group(2).strip()

                parsed = _parse_instalment_entry(payee, lines, txn_date, source)
                if not parsed:
                    continue

                key = (parsed["merchant"].upper(), parsed["total_n"], source)
                if key not in plans or parsed["current_n"] > plans[key]["current_n"]:
                    remaining = parsed["total_n"] - parsed["current_n"]
                    plans[key] = {
                        "merchant":       parsed["merchant"],
                        "total_n":        parsed["total_n"],
                        "current_n":      parsed["current_n"],
                        "monthly_amount": parsed["amount"],
                        "last_charge":    parsed["date"],
                        "remaining":      remaining,
                        "end_date":       _add_months(parsed["date"], remaining),
                        "source":         source.upper(),
                    }

    return sorted(plans.values(), key=lambda x: x["monthly_amount"], reverse=True)


# ── Excel helpers ─────────────────────────────────────────────────────────────

def header_style(cell):
    cell.font = Font(bold=True, color=C_HEADER_FG, name="Calibri", size=11)
    cell.fill = PatternFill("solid", fgColor=C_HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def subheader_style(cell):
    cell.font = Font(bold=True, name="Calibri", size=11)
    cell.fill = PatternFill("solid", fgColor=C_SUBHEADER_BG)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def total_style(cell):
    cell.font = Font(bold=True, name="Calibri", size=11)
    cell.fill = PatternFill("solid", fgColor=C_TOTAL_BG)


def set_col_widths(ws, widths: dict):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


# ── Report builder ────────────────────────────────────────────────────────────

def generate_cc_report(projection_months: int = 6) -> Path:
    print("Generating credit card commitment report ...")

    txns         = get_cc_expense_postings(["hsbc/*.journal", "ambcc/*.journal"])
    balance_hsbc = get_card_balance("Liabilities:CreditCard:HSBC")
    balance_amcc = get_card_balance("Liabilities:CreditCard:AmBankCC")
    recurring    = detect_recurring(txns)
    installments = extract_installments()
    proj_months  = project_months(projection_months)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Current Status ───────────────────────────────────────────────
    ws = wb.create_sheet("Current Status")
    ws.sheet_properties.tabColor = "1F3864"

    ws.merge_cells("A1:C1")
    t = ws["A1"]
    t.value = f"Credit Card Commitment Overview  (as at {date.today().strftime('%d %b %Y')})"
    t.font  = Font(bold=True, size=14, name="Calibri", color=C_HEADER_FG)
    t.fill  = PatternFill("solid", fgColor=C_HEADER_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    subs           = [r for r in recurring if r["is_subscription"]]
    regular        = [r for r in recurring if not r["is_subscription"]]
    highly_regular = [r for r in regular if r["months_seen"] >= 3]
    total_subs        = sum(r["avg_amount"] for r in subs)
    total_regular     = sum(r["avg_amount"] for r in highly_regular)
    total_instalment  = sum(p["monthly_amount"] for p in installments)
    total_monthly_committed = total_subs + total_regular + total_instalment

    data_months = sorted({(t["date"].year, t["date"].month) for t in txns})
    coverage_str = (
        f"{date(data_months[0][0], data_months[0][1], 1).strftime('%b %Y')} – "
        f"{date(data_months[-1][0], data_months[-1][1], 1).strftime('%b %Y')}"
        if data_months else "N/A"
    )

    status_rows = [
        ("HSBC Credit Card Balance",            balance_hsbc,                MYR_FORMAT),
        ("AmBank CC Balance (CARz S)",          balance_amcc,                MYR_FORMAT),
        ("Total Outstanding",                   balance_hsbc + balance_amcc, MYR_FORMAT),
        ("",                                    "",                           None),
        ("Months of Data",                      len(data_months),             "0"),
        ("Data Coverage",                       coverage_str,                 None),
        ("",                                    "",                           None),
        ("Fixed Subscriptions",                 len(subs),                    "0"),
        ("  → Est. Monthly (subscriptions)",    total_subs,                   MYR_FORMAT),
        ("Highly Regular Spend (3+ months)",    len(highly_regular),          "0"),
        ("  → Est. Monthly (regular)",          total_regular,                MYR_FORMAT),
        ("Instalment Plans Active",             len(installments),            "0"),
        ("  → Est. Monthly (instalments)",      total_instalment,             MYR_FORMAT),
        ("",                                    "",                           None),
        ("Est. Monthly Committed Total",        total_monthly_committed,      MYR_FORMAT),
        ("Est. Annual Committed Total",         total_monthly_committed * 12, MYR_FORMAT),
    ]

    bold_labels = {"Total Outstanding", "Est. Monthly Committed Total"}
    for r, (label, value, fmt) in enumerate(status_rows, start=2):
        lc = ws.cell(row=r, column=1, value=label)
        vc = ws.cell(row=r, column=2, value=value if value != "" else "")
        lc.font = Font(name="Calibri", size=11)
        if fmt == MYR_FORMAT and isinstance(value, float):
            vc.number_format = MYR_FORMAT
        elif fmt == "0" and isinstance(value, (int, float)):
            vc.number_format = "0"
        if label in bold_labels:
            lc.font = Font(bold=True, name="Calibri", size=11)
            vc.font = Font(bold=True, name="Calibri", size=11)
            vc.fill = PatternFill("solid", fgColor=C_TOTAL_BG)

    set_col_widths(ws, {"A": 38, "B": 20, "C": 5})

    # ── Sheet 2: Recurring Items (two sections) ───────────────────────────────
    ws2 = wb.create_sheet("Recurring Items")
    ws2.sheet_properties.tabColor = "2E7D32"

    col_headers = ["Merchant", "Category", "Avg Amount (MYR)", "Billing Day", "Months Seen", "Last Charged"]

    subscriptions  = [x for x in recurring if x["is_subscription"]]
    regular_spend  = [x for x in recurring if not x["is_subscription"]]

    def write_section(ws, start_row: int, section_label: str, items: list, bg_color: str) -> int:
        """Write a section header, column headers, data rows, subtotal. Returns next free row."""
        # Section label
        ws.merge_cells(f"A{start_row}:F{start_row}")
        sc = ws.cell(row=start_row, column=1, value=section_label)
        sc.font  = Font(bold=True, name="Calibri", size=11, color="FFFFFF")
        sc.fill  = PatternFill("solid", fgColor=bg_color)
        sc.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[start_row].height = 20
        start_row += 1

        # Column headers
        for c, h in enumerate(col_headers, 1):
            header_style(ws.cell(row=start_row, column=c, value=h))
        ws.row_dimensions[start_row].height = 22
        data_start = start_row + 1
        start_row += 1

        # Data rows
        for item in items:
            ws.cell(row=start_row, column=1, value=item["merchant"])
            ws.cell(row=start_row, column=2, value=item["category"])
            amt = ws.cell(row=start_row, column=3, value=item["avg_amount"])
            amt.number_format = MYR_FORMAT
            ws.cell(row=start_row, column=4, value=item["billing_day"])
            ws.cell(row=start_row, column=5, value=item["months_seen"])
            ws.cell(row=start_row, column=6, value=item["last_charge"].strftime("%Y-%m-%d"))
            if start_row % 2 == 0:
                for c in range(1, 7):
                    ws.cell(row=start_row, column=c).fill = PatternFill("solid", fgColor=C_ALT_ROW)
            start_row += 1

        # Subtotal row
        subtotal_r = start_row
        ws.cell(row=subtotal_r, column=1, value="Subtotal")
        sub_amt = ws.cell(row=subtotal_r, column=3,
                          value=f"=SUM(C{data_start}:C{subtotal_r - 1})")
        sub_amt.number_format = MYR_FORMAT
        for c in [1, 3]:
            total_style(ws.cell(row=subtotal_r, column=c))
        return subtotal_r + 2   # leave a blank row before next section

    row_cursor = 1
    sub1_row = None
    sub2_row = None

    if subscriptions:
        sub1_row = row_cursor + len(subscriptions) + 2  # will be the subtotal row
        row_cursor = write_section(ws2, row_cursor, "FIXED SUBSCRIPTIONS", subscriptions, "2E7D32")
        sub1_row = row_cursor - 2  # subtotal is 2 rows back (blank gap added)

    if regular_spend:
        sub2_row = row_cursor + len(regular_spend) + 2
        row_cursor = write_section(ws2, row_cursor, "REGULAR SPEND PATTERNS", regular_spend, "E65100")
        sub2_row = row_cursor - 2

    # Grand total
    grand_r = row_cursor
    ws2.cell(row=grand_r, column=1, value="TOTAL MONTHLY COMMITTED")
    refs = []
    if sub1_row:
        refs.append(f"C{sub1_row}")
    if sub2_row:
        refs.append(f"C{sub2_row}")
    grand_amt = ws2.cell(row=grand_r, column=3,
                         value=f"={'+'.join(refs)}" if refs else 0)
    grand_amt.number_format = MYR_FORMAT
    for c in [1, 3]:
        cell = ws2.cell(row=grand_r, column=c)
        cell.font = Font(bold=True, name="Calibri", size=12)
        cell.fill = PatternFill("solid", fgColor=C_TOTAL_BG)

    set_col_widths(ws2, {"A": 42, "B": 30, "C": 18, "D": 14, "E": 14, "F": 16})
    ws2.freeze_panes = ws2["B2"]

    # ── Sheet 3: Instalment Plans ─────────────────────────────────────────────
    ws3 = wb.create_sheet("Instalment Plans")
    ws3.sheet_properties.tabColor = "B71C1C"

    inst_headers = ["Merchant", "Card", "Monthly (MYR)", "Progress", "Total Months",
                    "Remaining", "End Date", "Total Remaining (MYR)"]
    for c, h in enumerate(inst_headers, 1):
        header_style(ws3.cell(row=1, column=c, value=h))
    ws3.row_dimensions[1].height = 22

    for r, plan in enumerate(installments, start=2):
        ws3.cell(row=r, column=1, value=plan["merchant"])
        ws3.cell(row=r, column=2, value=plan.get("source", ""))
        amt = ws3.cell(row=r, column=3, value=plan["monthly_amount"])
        amt.number_format = MYR_FORMAT
        ws3.cell(row=r, column=4, value=f"{plan['current_n']}/{plan['total_n']}")
        ws3.cell(row=r, column=5, value=plan["total_n"])
        ws3.cell(row=r, column=6, value=plan["remaining"])
        ws3.cell(row=r, column=7, value=plan["end_date"].strftime("%b %Y"))
        total_rem = ws3.cell(row=r, column=8, value=f"=C{r}*F{r}")
        total_rem.number_format = MYR_FORMAT
        if r % 2 == 0:
            for c in range(1, 9):
                ws3.cell(row=r, column=c).fill = PatternFill("solid", fgColor=C_ALT_ROW)

    # Totals row
    inst_tot_r = len(installments) + 2
    ws3.cell(row=inst_tot_r, column=1, value="TOTAL")
    inst_monthly = ws3.cell(row=inst_tot_r, column=3,
                            value=f"=SUM(C2:C{inst_tot_r-1})")
    inst_monthly.number_format = MYR_FORMAT
    inst_total_rem = ws3.cell(row=inst_tot_r, column=8,
                              value=f"=SUM(H2:H{inst_tot_r-1})")
    inst_total_rem.number_format = MYR_FORMAT
    for c in [1, 3, 8]:
        total_style(ws3.cell(row=inst_tot_r, column=c))

    set_col_widths(ws3, {"A": 42, "B": 10, "C": 16, "D": 12, "E": 14, "F": 12, "G": 12, "H": 22})
    ws3.freeze_panes = ws3["C2"]

    # ── Sheet 4: 6-Month Projection ───────────────────────────────────────────
    ws4 = wb.create_sheet(f"{projection_months}-Month Projection")
    ws4.sheet_properties.tabColor = "E65100"

    month_labels  = [m.strftime("%b %Y") for m in proj_months]
    proj_headers  = ["Merchant", "Category", "Billing Day"] + month_labels + ["Avg/Month"]
    n_month_cols  = len(proj_months)
    total_col     = len(proj_headers)
    first_mc_col  = get_column_letter(4)
    last_mc_col   = get_column_letter(3 + n_month_cols)

    for c, h in enumerate(proj_headers, 1):
        header_style(ws4.cell(row=1, column=c, value=h))
    ws4.row_dimensions[1].height = 22

    # Helper: write a projection section, return (start_row_of_data, subtotal_row, next_free_row)
    def write_proj_section(ws, start_r: int, label: str, color: str, items: list,
                           get_month_val) -> tuple[int, int]:
        # Section label
        ws.merge_cells(f"A{start_r}:{get_column_letter(total_col)}{start_r}")
        sc = ws.cell(row=start_r, column=1, value=label)
        sc.font  = Font(bold=True, name="Calibri", size=11, color="FFFFFF")
        sc.fill  = PatternFill("solid", fgColor=color)
        sc.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[start_r].height = 20
        data_start = start_r + 1

        for r, item in enumerate(items, start=data_start):
            ws.cell(row=r, column=1, value=item.get("merchant", ""))
            ws.cell(row=r, column=2, value=item.get("category", ""))
            ws.cell(row=r, column=3, value=item.get("billing_day", "~3"))
            for mi, pm in enumerate(proj_months):
                col = 4 + mi
                val = get_month_val(item, mi)
                if val is not None:
                    cell = ws.cell(row=r, column=col, value=val)
                    cell.number_format = MYR_FORMAT
            avg_cell = ws.cell(row=r, column=total_col,
                               value=f"=IFERROR(AVERAGE({first_mc_col}{r}:{last_mc_col}{r}),\"\")")
            avg_cell.number_format = MYR_FORMAT
            if r % 2 == 0:
                for c in range(1, total_col + 1):
                    ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=C_ALT_ROW)

        sub_r = data_start + len(items)
        ws.cell(row=sub_r, column=1, value="Subtotal")
        for mc in range(4, total_col + 1):
            cl = get_column_letter(mc)
            cell = ws.cell(row=sub_r, column=mc,
                           value=f"=SUM({cl}{data_start}:{cl}{sub_r-1})")
            cell.number_format = MYR_FORMAT
            total_style(cell)
        total_style(ws.cell(row=sub_r, column=1))
        return sub_r, sub_r + 2   # subtotal row, next cursor

    sub_rows = []
    row_cursor = 2

    # Section 1: Fixed Subscriptions
    if subs:
        sub_r, row_cursor = write_proj_section(
            ws4, row_cursor, "FIXED SUBSCRIPTIONS", "2E7D32", subs,
            lambda item, mi: item["avg_amount"]
        )
        sub_rows.append(sub_r)

    # Section 2: Highly Regular Spend (3+ months)
    if highly_regular:
        sub_r, row_cursor = write_proj_section(
            ws4, row_cursor, "HIGHLY REGULAR SPEND (3+ months)", "1565A0", highly_regular,
            lambda item, mi: item["avg_amount"]
        )
        sub_rows.append(sub_r)

    # Section 3: Instalment Plans (cap at remaining months)
    if installments:
        def instalment_month_val(plan, month_index):
            # month_index 0 = next month; show amount only while instalments remain
            return plan["monthly_amount"] if month_index < plan["remaining"] else None

        inst_items = [{**p, "category": "Instalment", "billing_day": "~3"} for p in installments]
        sub_r, row_cursor = write_proj_section(
            ws4, row_cursor, "INSTALMENT PLANS", "B71C1C", inst_items,
            instalment_month_val
        )
        sub_rows.append(sub_r)

    # Grand total row
    grand_r = row_cursor
    ws4.cell(row=grand_r, column=1, value="GRAND TOTAL")
    for mc in range(4, total_col + 1):
        cl = get_column_letter(mc)
        refs = "+".join(f"{cl}{sr}" for sr in sub_rows)
        cell = ws4.cell(row=grand_r, column=mc, value=f"={refs}")
        cell.number_format = MYR_FORMAT
        cell.font = Font(bold=True, name="Calibri", size=12)
        cell.fill = PatternFill("solid", fgColor=C_TOTAL_BG)
    ws4.cell(row=grand_r, column=1).font = Font(bold=True, name="Calibri", size=12)
    ws4.cell(row=grand_r, column=1).fill = PatternFill("solid", fgColor=C_TOTAL_BG)

    set_col_widths(ws4, {"A": 42, "B": 28, "C": 12})
    for mc in range(4, total_col + 1):
        ws4.column_dimensions[get_column_letter(mc)].width = 14
    ws4.freeze_panes = "D2"

    # ── Sheet 5: Transaction History ──────────────────────────────────────────
    ws5 = wb.create_sheet("Transaction History")
    ws5.sheet_properties.tabColor = "4A148C"

    txn_headers = ["Date", "Card", "Merchant", "Category", "Amount (MYR)"]
    for c, h in enumerate(txn_headers, 1):
        header_style(ws5.cell(row=1, column=c, value=h))
    ws5.row_dimensions[1].height = 22

    for r, t in enumerate(txns, start=2):
        ws5.cell(row=r, column=1, value=t["date"].strftime("%Y-%m-%d"))
        ws5.cell(row=r, column=2, value=t.get("source", "").upper())
        ws5.cell(row=r, column=3, value=t["merchant"])
        ws5.cell(row=r, column=4, value=t["category"].replace("Expenses:", ""))
        amt = ws5.cell(row=r, column=5, value=t["amount"])
        amt.number_format = MYR_FORMAT
        if r % 2 == 0:
            for c in range(1, 6):
                ws5.cell(row=r, column=c).fill = PatternFill("solid", fgColor=C_ALT_ROW)

    set_col_widths(ws5, {"A": 12, "B": 10, "C": 48, "D": 32, "E": 16})
    ws5.freeze_panes = ws5["C2"]

    # ── Save ──────────────────────────────────────────────────────────────────
    config.REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    out_path = config.REPORTS_DIR / "cc_commitment.xlsx"
    wb.save(out_path)
    print(f"  Instalment plans found   : {len(installments)}")
    print(f"  Recurring items detected : {len(recurring)}")
    print(f"  Est. monthly committed   : MYR {total_monthly_committed:,.2f}")
    print(f"  Saved → {out_path}")
    return out_path


# ── CLI ───────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate HSBC credit card commitment report")
    parser.add_argument("--months", type=int, default=6,
                        help="Number of months to project forward (default: 6)")
    args = parser.parse_args()
    generate_cc_report(projection_months=args.months)
