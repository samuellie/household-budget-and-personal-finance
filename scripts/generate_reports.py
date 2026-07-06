"""
Excel Report Generator
========================
Queries ledger CLI and produces formatted Excel workbooks.

Reports:
  - Monthly summary: summary, category breakdown, full transaction list
  - Annual summary: P&L statement, monthly trends, category totals, account balances

Usage:
    python3 generate_reports.py --month 2025-04
    python3 generate_reports.py --month 2025-04 --month 2025-05
    python3 generate_reports.py --annual
    python3 generate_reports.py --all         # generate everything
"""

import argparse
import re
import subprocess
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
import config

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip3 install --break-system-packages openpyxl")
    sys.exit(1)

# ── Base colours ──────────────────────────────────────────────────────────────
C_HEADER_BG    = "1F3864"
C_HEADER_FG    = "FFFFFF"
C_INCOME_BG    = "E8F5E9"
C_EXPENSE_BG   = "FFEBEE"
C_SUBHEADER_BG = "D6E4F7"
C_ALT_ROW      = "F5F5F5"
C_TOTAL_BG     = "FFF9C4"

# ── P&L section colours ───────────────────────────────────────────────────────
C_INC_HDR  = "1B5E20"   # dark green
C_FIX_HDR  = "BF360C"   # dark burnt orange
C_VAR_HDR  = "4A148C"   # dark purple
C_INC_ROW  = "E8F5E9"   # light green
C_FIX_ROW  = "FBE9E7"   # light orange
C_VAR_ROW  = "EDE7F6"   # light lavender
C_INC_SUB  = "A5D6A7"   # medium green
C_FIX_SUB  = "FFCC80"   # amber
C_VAR_SUB  = "CE93D8"   # medium purple
C_GRAND_EXP = "EF9A9A"  # red — total expenses row
C_GRAND_NET = "FFF176"  # yellow — net savings row

MYR_FORMAT = '#,##0.00'

# ── P&L structure definition ──────────────────────────────────────────────────
# Each section: name, colours, and line items with their ledger account prefixes.
# Prefix matching is hierarchical: "Expenses:Subscriptions" also catches
# Expenses:Subscriptions:Spotify, Expenses:Subscriptions:Google, etc.

PNL_STRUCTURE = [
    {
        "name": "INCOME",
        "hdr": C_INC_HDR, "row": C_INC_ROW, "sub": C_INC_SUB,
        "lines": [
            ("Samuel's Salary",                  ["Income:Salary:Samuel"]),
            ("Fui Yee's Salary",                 ["Income:Salary:FuiYee"]),
            ("Interest",                          ["Income:Interest"]),
            ("Freelance",                          ["Income:Freelance"]),
            ("Refunds & Other",                   ["Income:Refund", "Income:Rental", "Income:Other"]),
        ],
    },
    {
        "name": "COMMITTED EXPENSES",
        "hdr": C_FIX_HDR, "row": C_FIX_ROW, "sub": C_FIX_SUB,
        "lines": [
            ("Mortgage — Fui Yee CIMB (RM 3,551/mth)", ["Expenses:Housing:Mortgage"]),
            ("Rent",                              ["Expenses:Housing:Rent"]),
            ("Utilities",                         ["Expenses:Utilities"]),
            ("Insurance",                         ["Expenses:Insurance"]),
            ("Subscriptions",                     ["Expenses:Subscriptions"]),
            ("Instalments",                       ["Expenses:Instalment"]),
            ("Gym",                               ["Expenses:Healthcare:Gym"]),
        ],
    },
    {
        "name": "VARIABLE EXPENSES",
        "hdr": C_VAR_HDR, "row": C_VAR_ROW, "sub": C_VAR_SUB,
        "lines": [
            ("Groceries",                         ["Expenses:Groceries"]),
            ("Dining Out & Coffee",               ["Expenses:Dining", "Expenses:Coffee"]),
            ("Transport",                         ["Expenses:Transport", "Expenses:Petrol",
                                                   "Expenses:Parking"]),
            ("Shopping",                          ["Expenses:Shopping"]),
            ("Personal & Sports",                 ["Expenses:Personal", "Expenses:Sports"]),
            ("Entertainment",                     ["Expenses:Entertainment"]),
            ("Healthcare",                        ["Expenses:Healthcare:Medical",
                                                   "Expenses:Healthcare:Pharmacy",
                                                   "Expenses:Health"]),
            ("Housing Maintenance",               ["Expenses:Housing:Maintenance"]),
            ("Gifts & Education",                 ["Expenses:Gifts", "Expenses:Education"]),
            ("Bank Charges",                      ["Expenses:Bank"]),
            ("Tax",                               ["Expenses:Tax"]),
            ("Business",                          ["Expenses:Business"]),
            ("Divo House",                        ["Expenses:Divo"]),
            ("Uncategorized",                     ["Expenses:Uncategorized"]),
        ],
    },
]


# ── Ledger queries ────────────────────────────────────────────────────────────

def run_ledger(args: list) -> str:
    cmd = ["ledger", "-f", str(config.MAIN_JOURNAL)] + args
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode != 0:
        raise RuntimeError(f"ledger error: {result.stderr.strip()}")
    return result.stdout


def ledger_balance(account_pattern: str, period: str | None = None, end: str | None = None) -> dict:
    args = ["bal", account_pattern, "--flat", "--no-total", "--exchange", "MYR"]
    if period:
        args += ["--period", period]
    if end:
        args += ["--end", end]
    try:
        output = run_ledger(args)
    except RuntimeError:
        return {}
    result = {}
    for line in output.splitlines():
        m = re.match(r"\s*(?:MYR\s+)?([-\d,]+\.\d{2})\s{2,}(.+)", line)
        if m:
            amount = float(m.group(1).replace(",", ""))
            account = m.group(2).strip()
            result[account] = amount
    return result


def ledger_register_csv(account_pattern: str, period: str) -> list:
    args = ["csv", "--period", period]
    if account_pattern:
        args.append(account_pattern)
    try:
        output = run_ledger(args)
    except RuntimeError:
        return []
    import csv, io
    rows = []
    reader = csv.reader(io.StringIO(output))
    for row in reader:
        if len(row) >= 7:
            try:
                rows.append({
                    "date":    row[0].strip('"'),
                    "payee":   row[2].strip('"'),
                    "account": row[3].strip('"'),
                    "amount":  float(row[5].strip('"').replace(",", "")),
                })
            except (ValueError, IndexError):
                continue
    return rows


# ── Excel helpers ─────────────────────────────────────────────────────────────

def header_style(cell, bold=True):
    cell.font = Font(bold=bold, color=C_HEADER_FG, name="Calibri", size=11)
    cell.fill = PatternFill("solid", fgColor=C_HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def subheader_style(cell):
    cell.font = Font(bold=True, name="Calibri", size=11)
    cell.fill = PatternFill("solid", fgColor=C_SUBHEADER_BG)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def total_style(cell, is_income=False, is_expense=False):
    cell.font = Font(bold=True, name="Calibri", size=11)
    bg = C_TOTAL_BG
    if is_income:
        bg = "A5D6A7"
    elif is_expense:
        bg = "EF9A9A"
    cell.fill = PatternFill("solid", fgColor=bg)


def set_col_widths(ws, widths: dict):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def freeze(ws, cell="B2"):
    ws.freeze_panes = ws[cell]


# ── P&L helpers ───────────────────────────────────────────────────────────────

# ── Person scope ──────────────────────────────────────────────────────────────
# Accounts are person-first: <Root>:<Person>:<...>. Reports run in three scopes:
#   None      → Combined household (both people aggregated under neutral categories)
#   "Samuel"  → Samuel only
#   "FuiYee"  → Fui Yee only
SCOPE: str | None = None
PERSON_SEG = ("Samuel", "FuiYee")


def scope_root(root: str) -> str:
    """Ledger query pattern for a top-level root under the current scope."""
    return f"{root}:{SCOPE}" if SCOPE else root


def strip_person(account: str) -> str:
    """Remove the person segment so person-neutral prefixes/labels match:
    Expenses:Samuel:Groceries → Expenses:Groceries ; Equity:Transfer unchanged."""
    parts = account.split(":")
    if len(parts) >= 2 and parts[1] in PERSON_SEG:
        return ":".join([parts[0]] + parts[2:])
    return account


def scope_label() -> str:
    return {None: "Combined Household", "Samuel": "Samuel", "FuiYee": "Fui Yee"}[SCOPE]


def scope_suffix() -> str:
    return {None: "combined", "Samuel": "samuel", "FuiYee": "fuiyee"}[SCOPE]


def _acc_matches(account: str, prefixes: list) -> bool:
    account = strip_person(account)
    return any(account == p or account.startswith(p + ":") for p in prefixes)


def _fetch_period_data(all_months: list) -> dict:
    """Return {period: {account: float}} with income negated to positive."""
    data = {}
    for y, m in all_months:
        period = f"{y}-{m:02d}"
        pd = {}
        for acc, amt in ledger_balance(scope_root("Income"),   period=period).items():
            pd[acc] = -amt          # ledger income is negative → make positive
        for acc, amt in ledger_balance(scope_root("Expenses"), period=period).items():
            if amt > 0:
                pd[acc] = amt
        data[period] = pd
    return data


# ── P&L Annual Sheet ──────────────────────────────────────────────────────────

def build_pnl_annual_sheet(wb, all_months: list, month_labels: list) -> None:
    """Build 'P&L Statement' as the first sheet — company-style income statement."""
    ws = wb.create_sheet("P&L Statement", 0)
    ws.sheet_properties.tabColor = "1F3864"

    n_months  = len(month_labels)
    total_col = n_months + 2                    # col1=label, 2..n+1=months, last=total
    last_data = get_column_letter(n_months + 1)

    pdata = _fetch_period_data(all_months)

    def row_vals(prefixes):
        return [
            sum(v for a, v in pdata.get(f"{y}-{m:02d}", {}).items()
                if _acc_matches(a, prefixes))
            for y, m in all_months
        ]

    def bg(color):
        return PatternFill("solid", fgColor=color)

    # ── Title ─────────────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(total_col)}1")
    t = ws["A1"]
    t.value = f"Annual P&L Statement  —  {scope_label()}  (FY 2025 / 2026)"
    t.font  = Font(bold=True, size=14, name="Calibri", color="FFFFFF")
    t.fill  = bg(C_HEADER_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # ── Column headers ────────────────────────────────────────────────────────
    header_style(ws.cell(row=2, column=1, value=""))
    for c, lbl in enumerate(month_labels, start=2):
        yy, mm = lbl.split("-")
        short = datetime(int(yy), int(mm), 1).strftime("%b '%y")
        header_style(ws.cell(row=2, column=c, value=short))
    header_style(ws.cell(row=2, column=total_col, value="ANNUAL TOTAL"))
    ws.row_dimensions[2].height = 24

    cur = 3
    subtotals: dict[str, int] = {}

    for sec in PNL_STRUCTURE:
        # Section header
        ws.merge_cells(f"A{cur}:{get_column_letter(total_col)}{cur}")
        h = ws.cell(row=cur, column=1, value=f"  {sec['name']}")
        h.font  = Font(bold=True, size=11, name="Calibri", color="FFFFFF")
        h.fill  = bg(sec["hdr"])
        h.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[cur].height = 20
        cur += 1

        live_rows: list[int] = []
        for label, prefixes in sec["lines"]:
            vals = row_vals(prefixes)
            if all(v == 0 for v in vals):
                continue                        # omit lines with no data

            ws.cell(row=cur, column=1, value=f"   {label}").fill = bg(sec["row"])
            ws.cell(row=cur, column=1).font = Font(name="Calibri", size=10)

            for c, v in enumerate(vals, start=2):
                cell = ws.cell(row=cur, column=c, value=v if v else None)
                cell.fill = bg(sec["row"])
                cell.font = Font(name="Calibri", size=10)
                if v:
                    cell.number_format = MYR_FORMAT

            tc = ws.cell(row=cur, column=total_col,
                         value=f"=SUM(B{cur}:{last_data}{cur})")
            tc.number_format = MYR_FORMAT
            tc.fill = bg(sec["row"])
            tc.font = Font(bold=True, name="Calibri", size=10)

            live_rows.append(cur)
            cur += 1

        if live_rows:
            sc = ws.cell(row=cur, column=1, value=f"TOTAL  {sec['name']}")
            sc.font = Font(bold=True, size=11, name="Calibri")
            sc.fill = bg(sec["sub"])
            for c in range(2, total_col + 1):
                cl = get_column_letter(c)
                cell = ws.cell(row=cur, column=c,
                               value="=" + "+".join(f"{cl}{r}" for r in live_rows))
                cell.number_format = MYR_FORMAT
                cell.font = Font(bold=True, size=11, name="Calibri")
                cell.fill = bg(sec["sub"])
            subtotals[sec["name"]] = cur
            ws.row_dimensions[cur].height = 20
            cur += 2    # blank gap between sections

    # ── Grand totals ──────────────────────────────────────────────────────────
    inc_r = subtotals.get("INCOME")
    fix_r = subtotals.get("COMMITTED EXPENSES")
    var_r = subtotals.get("VARIABLE EXPENSES")

    # Total Expenses = Committed + Variable
    exp_parts = [r for r in (fix_r, var_r) if r]
    if exp_parts:
        ws.cell(row=cur, column=1, value="TOTAL EXPENSES").fill = bg(C_GRAND_EXP)
        ws.cell(row=cur, column=1).font = Font(bold=True, size=11, name="Calibri")
        for c in range(2, total_col + 1):
            cl = get_column_letter(c)
            cell = ws.cell(row=cur, column=c,
                           value="=" + "+".join(f"{cl}{r}" for r in exp_parts))
            cell.number_format = MYR_FORMAT
            cell.font = Font(bold=True, size=11, name="Calibri")
            cell.fill = bg(C_GRAND_EXP)
        total_exp_row = cur
        ws.row_dimensions[cur].height = 20
        cur += 1
    else:
        total_exp_row = None

    # Net Savings
    if inc_r and total_exp_row:
        net_row = cur
        ws.cell(row=net_row, column=1, value="NET SAVINGS").fill = bg(C_GRAND_NET)
        ws.cell(row=net_row, column=1).font = Font(bold=True, size=13, name="Calibri")
        for c in range(2, total_col + 1):
            cl = get_column_letter(c)
            cell = ws.cell(row=net_row, column=c,
                           value=f"={cl}{inc_r}-{cl}{total_exp_row}")
            cell.number_format = MYR_FORMAT
            cell.font = Font(bold=True, size=13, name="Calibri")
            cell.fill = bg(C_GRAND_NET)
        ws.row_dimensions[net_row].height = 24
        cur += 1

        # Savings Rate %
        ws.cell(row=cur, column=1, value="  Savings Rate").font = Font(
            name="Calibri", size=10, italic=True)
        for c in range(2, total_col + 1):
            cl = get_column_letter(c)
            cell = ws.cell(row=cur, column=c,
                           value=f"=IF({cl}{inc_r}=0,0,{cl}{net_row}/{cl}{inc_r})")
            cell.number_format = "0.0%"
            cell.font = Font(name="Calibri", size=10, italic=True)
        cur += 1

        # Committed Expense Ratio %
        if fix_r:
            ws.cell(row=cur, column=1, value="  Committed Expense Ratio").font = Font(
                name="Calibri", size=10, italic=True)
            for c in range(2, total_col + 1):
                cl = get_column_letter(c)
                cell = ws.cell(row=cur, column=c,
                               value=f"=IF({cl}{inc_r}=0,0,{cl}{fix_r}/{cl}{inc_r})")
                cell.number_format = "0.0%"
                cell.font = Font(name="Calibri", size=10, italic=True)

    # ── Column widths & freeze ────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 42
    for c in range(2, total_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 13
    ws.freeze_panes = "B3"


# ── Monthly Report ────────────────────────────────────────────────────────────

def generate_monthly_report(year: int, month: int) -> Path:
    period = f"{year}-{month:02d}"
    print(f"Generating monthly report for {period} ...")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary")
    ws_sum.sheet_properties.tabColor = "1F3864"

    income_bal  = ledger_balance(scope_root("Income"), period=period)
    expense_bal = ledger_balance(scope_root("Expenses"), period=period)
    asset_bal   = ledger_balance(scope_root("Assets"), end=f"{year}-{month:02d}-01")

    total_income   = -sum(income_bal.values())
    total_expenses =  sum(v for v in expense_bal.values() if v > 0)
    net_savings    = total_income - total_expenses
    savings_rate   = (net_savings / total_income * 100) if total_income else 0

    ws_sum.merge_cells("A1:D1")
    title = ws_sum["A1"]
    title.value = f"Monthly Financial Summary — {scope_label()} — {datetime(year, month, 1).strftime('%B %Y')}"
    title.font = Font(bold=True, size=14, name="Calibri", color=C_HEADER_FG)
    title.fill = PatternFill("solid", fgColor=C_HEADER_BG)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 30

    rows = [
        ("", ""),
        ("INCOME / CASH FLOW", ""),
        ("Total Income",   total_income),
        ("Total Expenses", total_expenses),
        ("Net Savings",    "=B4-B5"),
        ("Savings Rate",   "=IF(B4=0,0,B6/B4)"),
        ("", ""),
        ("ACCOUNT BALANCES (month-end)", ""),
    ]
    for acc, bal in sorted(asset_bal.items()):
        disp = strip_person(acc).replace("Assets:Bank:", "").replace("Assets:", "")
        rows.append((disp, bal))

    for r, (label, value) in enumerate(rows, start=2):
        cell_label = ws_sum.cell(row=r, column=1, value=label)
        cell_value = ws_sum.cell(row=r, column=2, value=value if value != "" else "")

        if label in ("INCOME / CASH FLOW", "ACCOUNT BALANCES (month-end)"):
            subheader_style(cell_label)
            ws_sum.merge_cells(f"A{r}:D{r}")
            continue

        if label == "Total Income":
            total_style(cell_value, is_income=True)
            cell_value.number_format = MYR_FORMAT
        elif label == "Total Expenses":
            total_style(cell_value, is_expense=True)
            cell_value.number_format = MYR_FORMAT
        elif label == "Net Savings":
            total_style(cell_value, is_income=(net_savings >= 0), is_expense=(net_savings < 0))
            cell_value.number_format = MYR_FORMAT
        elif label == "Savings Rate":
            cell_value.number_format = "0.0%"
        elif isinstance(value, float):
            cell_value.number_format = MYR_FORMAT

    set_col_widths(ws_sum, {"A": 35, "B": 18, "C": 5, "D": 5})

    # ── Sheet 2: Category Breakdown ───────────────────────────────────────────
    ws_cat = wb.create_sheet("Category Breakdown")
    ws_cat.sheet_properties.tabColor = "C62828"

    headers = ["Category", "Amount (MYR)", "% of Spending"]
    for c, h in enumerate(headers, 1):
        cell = ws_cat.cell(row=1, column=c, value=h)
        header_style(cell)
    ws_cat.row_dimensions[1].height = 22

    from collections import defaultdict
    _agg = defaultdict(float)
    for acc, amt in expense_bal.items():
        if amt > 0:
            _agg[strip_person(acc).replace("Expenses:", "")] += amt
    expense_items = sorted(_agg.items(), key=lambda x: x[1], reverse=True)

    for r, (category, amount) in enumerate(expense_items, start=2):
        ws_cat.cell(row=r, column=1, value=category)
        amt_cell = ws_cat.cell(row=r, column=2, value=amount)
        amt_cell.number_format = MYR_FORMAT
        pct_cell = ws_cat.cell(row=r, column=3, value=(amount / total_expenses) if total_expenses else 0)
        pct_cell.number_format = "0.0%"
        if r % 2 == 0:
            for c in range(1, 4):
                ws_cat.cell(row=r, column=c).fill = PatternFill("solid", fgColor=C_ALT_ROW)

    total_row = len(expense_items) + 2
    ws_cat.cell(row=total_row, column=1, value="TOTAL")
    total_amt = ws_cat.cell(row=total_row, column=2,
                            value=f"=SUM(B2:B{total_row - 1})")
    total_amt.number_format = MYR_FORMAT
    for c in [1, 2]:
        total_style(ws_cat.cell(row=total_row, column=c), is_expense=True)

    set_col_widths(ws_cat, {"A": 38, "B": 18, "C": 16})
    freeze(ws_cat)

    # ── Sheet 3: Transactions ─────────────────────────────────────────────────
    ws_txn = wb.create_sheet("Transactions")
    ws_txn.sheet_properties.tabColor = "2E7D32"

    txn_headers = ["Date", "Payee / Description", "Account", "Amount (MYR)"]
    for c, h in enumerate(txn_headers, 1):
        cell = ws_txn.cell(row=1, column=c, value=h)
        header_style(cell)
    ws_txn.row_dimensions[1].height = 22

    all_txns = ledger_register_csv(SCOPE or "", period=period)
    for r, txn in enumerate(all_txns, start=2):
        ws_txn.cell(row=r, column=1, value=txn["date"])
        ws_txn.cell(row=r, column=2, value=txn["payee"])
        ws_txn.cell(row=r, column=3, value=txn["account"])
        amt_cell = ws_txn.cell(row=r, column=4, value=txn["amount"])
        amt_cell.number_format = MYR_FORMAT
        if txn["amount"] > 0:
            amt_cell.font = Font(color="1B5E20")
        elif txn["amount"] < 0:
            amt_cell.font = Font(color="B71C1C")
        if r % 2 == 0:
            for c in range(1, 5):
                ws_txn.cell(row=r, column=c).fill = PatternFill("solid", fgColor=C_ALT_ROW)

    set_col_widths(ws_txn, {"A": 12, "B": 50, "C": 35, "D": 16})
    freeze(ws_txn)

    config.REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    out_path = config.REPORTS_DIR / f"monthly_{scope_suffix()}_{period}.xlsx"
    wb.save(out_path)
    print(f"  Saved → {out_path}")
    return out_path


# ── Annual Report ─────────────────────────────────────────────────────────────

def generate_annual_report(start_year: int = 2025, end_year: int = 2026) -> Path:
    print(f"Generating annual report ({start_year}–{end_year}) ...")

    # Only include months from first data (Mar 2025) up to current month
    all_months = [
        (y, m) for y in range(start_year, end_year + 1)
        for m in range(1, 13)
        if date(y, m, 1) >= date(2025, 3, 1)
        and date(y, m, 1) <= date.today().replace(day=1)
    ]

    month_labels = [f"{y}-{m:02d}" for y, m in all_months]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: P&L Statement (new — detailed income statement) ─────────────
    build_pnl_annual_sheet(wb, all_months, month_labels)

    # ── Sheet 2: Monthly Trends ───────────────────────────────────────────────
    ws_trend = wb.create_sheet("Monthly Trends")
    headers = ["Month", "Income", "Expenses", "Net Savings", "Savings Rate%"]
    for c, h in enumerate(headers, 1):
        cell = ws_trend.cell(row=1, column=c, value=h)
        header_style(cell)
    ws_trend.row_dimensions[1].height = 22

    trend_data = []
    for y, m in all_months:
        period = f"{y}-{m:02d}"
        inc_bal  = ledger_balance(scope_root("Income"),   period=period)
        exp_bal  = ledger_balance(scope_root("Expenses"), period=period)
        income   = -sum(inc_bal.values())
        expenses = sum(v for v in exp_bal.values() if v > 0)
        net      = income - expenses
        rate     = (net / income) if income else 0
        trend_data.append((f"{y}-{m:02d}", income, expenses, net, rate))

    for r, (month, income, expenses, net, rate) in enumerate(trend_data, start=2):
        ws_trend.cell(row=r, column=1, value=month)
        for c, val in enumerate([income, expenses, net], start=2):
            cell = ws_trend.cell(row=r, column=c, value=val)
            cell.number_format = MYR_FORMAT
        rate_cell = ws_trend.cell(row=r, column=5, value=f"=IF(B{r}=0,0,D{r}/B{r})")
        rate_cell.number_format = "0.0%"
        if r % 2 == 0:
            for c in range(1, 6):
                ws_trend.cell(row=r, column=c).fill = PatternFill("solid", fgColor=C_ALT_ROW)

    if trend_data:
        first_tr = 2
        last_tr  = len(trend_data) + 1
        tot_r    = last_tr + 1
        ws_trend.cell(row=tot_r, column=1, value="TOTAL")
        for c, col in enumerate(["B", "C", "D"], start=2):
            cell = ws_trend.cell(row=tot_r, column=c,
                                 value=f"=SUM({col}{first_tr}:{col}{last_tr})")
            cell.number_format = MYR_FORMAT
            total_style(cell)
        rate_cell = ws_trend.cell(row=tot_r, column=5,
                                  value=f"=IF(B{tot_r}=0,0,D{tot_r}/B{tot_r})")
        rate_cell.number_format = "0.0%"
        total_style(rate_cell)
        total_style(ws_trend.cell(row=tot_r, column=1))

    set_col_widths(ws_trend, {"A": 12, "B": 16, "C": 16, "D": 16, "E": 14})
    freeze(ws_trend)

    # ── Sheet 3: Category Annual ──────────────────────────────────────────────
    ws_cat = wb.create_sheet("Category Annual")

    all_categories = set()
    monthly_cats = {}
    for y, m in all_months:
        period = f"{y}-{m:02d}"
        bal = ledger_balance(scope_root("Expenses"), period=period)
        _cats = defaultdict(float)
        for k, v in bal.items():
            if v > 0:
                _cats[strip_person(k).replace("Expenses:", "")] += v
        monthly_cats[period] = dict(_cats)
        all_categories.update(monthly_cats[period].keys())

    sorted_cats = sorted(all_categories)

    ws_cat.cell(row=1, column=1, value="Category")
    header_style(ws_cat.cell(row=1, column=1))
    for c, lbl in enumerate(month_labels, start=2):
        cell = ws_cat.cell(row=1, column=c, value=lbl)
        header_style(cell)
    total_col = len(month_labels) + 2
    cell = ws_cat.cell(row=1, column=total_col, value="Total")
    header_style(cell)

    last_month_col = get_column_letter(total_col - 1)

    for r, cat in enumerate(sorted_cats, start=2):
        ws_cat.cell(row=r, column=1, value=cat)
        for c, (y, m) in enumerate(all_months, start=2):
            period = f"{y}-{m:02d}"
            val = monthly_cats[period].get(cat, 0)
            cell = ws_cat.cell(row=r, column=c, value=val if val else None)
            if val:
                cell.number_format = MYR_FORMAT
        total_cell = ws_cat.cell(row=r, column=total_col,
                                 value=f"=SUM(B{r}:{last_month_col}{r})")
        total_cell.number_format = MYR_FORMAT
        total_style(total_cell)
        if r % 2 == 0:
            for c in range(1, total_col + 1):
                ws_cat.cell(row=r, column=c).fill = PatternFill("solid", fgColor=C_ALT_ROW)

    total_row = len(sorted_cats) + 2
    last_data_row = total_row - 1
    ws_cat.cell(row=total_row, column=1, value="TOTAL")
    total_style(ws_cat.cell(row=total_row, column=1), is_expense=True)
    for c in range(2, total_col + 1):
        col_letter = get_column_letter(c)
        cell = ws_cat.cell(row=total_row, column=c,
                           value=f"=SUM({col_letter}2:{col_letter}{last_data_row})")
        cell.number_format = MYR_FORMAT
        total_style(cell, is_expense=True)

    ws_cat.column_dimensions["A"].width = 38
    for c in range(2, total_col + 1):
        ws_cat.column_dimensions[get_column_letter(c)].width = 14
    freeze(ws_cat)

    # ── Sheet 4: Account Balances ─────────────────────────────────────────────
    ws_bal = wb.create_sheet("Account Balances")

    balance_headers = ["Account"] + month_labels
    for c, h in enumerate(balance_headers, 1):
        cell = ws_bal.cell(row=1, column=c, value=h)
        header_style(cell)

    account_names = set()
    monthly_balances = {}
    all_periods = [f"{y}-{m:02d}" for y, m in all_months]
    for y, m in all_months:
        if m == 12:
            end_date = f"{y+1}-01-01"
        else:
            end_date = f"{y}-{m+1:02d}-01"
        bal = ledger_balance(f"{scope_root('Assets')}|{scope_root('Liabilities')}", end=end_date)
        period = f"{y}-{m:02d}"
        monthly_balances[period] = bal
        account_names.update(bal.keys())

    # Track first period each account appears so pre-existence periods stay blank
    account_first_period = {}
    for acc in account_names:
        for period in all_periods:
            if acc in monthly_balances[period]:
                account_first_period[acc] = period
                break

    for r, acc in enumerate(sorted(account_names), start=2):
        _a = strip_person(acc) if SCOPE else acc
        _disp = (_a.replace("Assets:", "").replace("Liabilities:", "")
                    .replace("Bank:", "").replace("CreditCard:", "CC: "))
        ws_bal.cell(row=r, column=1, value=_disp)
        first = account_first_period.get(acc, all_periods[-1])
        for c, (y, m) in enumerate(all_months, start=2):
            period = f"{y}-{m:02d}"
            if period < first:
                val = None  # account didn't exist yet
            else:
                val = monthly_balances[period].get(acc, 0)
            cell = ws_bal.cell(row=r, column=c, value=val)
            if val is not None:
                cell.number_format = MYR_FORMAT

    ws_bal.column_dimensions["A"].width = 25
    for c in range(2, len(all_months) + 2):
        ws_bal.column_dimensions[get_column_letter(c)].width = 14
    freeze(ws_bal)

    # ── Save ──────────────────────────────────────────────────────────────────
    config.REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    out_path = config.REPORTS_DIR / f"annual_{scope_suffix()}_{start_year}-{end_year}.xlsx"
    wb.save(out_path)
    print(f"  Saved → {out_path}")
    return out_path


# ── CLI ───────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate Excel financial reports from ledger data")
    parser.add_argument("--month", action="append", metavar="YYYY-MM",
                        help="Generate monthly report (can specify multiple)")
    parser.add_argument("--annual", action="store_true", help="Generate annual summary report")
    parser.add_argument("--all", action="store_true", help="Generate all monthly + annual reports")
    parser.add_argument("--scope", choices=["combined", "samuel", "fuiyee", "all"],
                        default="all",
                        help="Person scope: combined, samuel, fuiyee, or all (default: all three)")
    args = parser.parse_args()

    if not any([args.month, args.annual, args.all]):
        parser.print_help()
        sys.exit(0)

    scope_map = {"combined": None, "samuel": "Samuel", "fuiyee": "FuiYee"}
    scopes = [None, "Samuel", "FuiYee"] if args.scope == "all" else [scope_map[args.scope]]

    months_to_run = []
    if args.all:
        months_to_run = (
            [(2025, m) for m in range(3, 13)] +
            [(2026, m) for m in range(1, 7)]
        )
    elif args.month:
        for m_str in args.month:
            y, m = int(m_str[:4]), int(m_str[5:7])
            months_to_run.append((y, m))

    for sc in scopes:
        SCOPE = sc
        print(f"\n########## SCOPE: {scope_label()} ##########")
        if args.all or args.annual:
            generate_annual_report()
        for y, m in months_to_run:
            try:
                generate_monthly_report(y, m)
            except RuntimeError as e:
                print(f"  SKIP {y}-{m:02d}: {e}")
